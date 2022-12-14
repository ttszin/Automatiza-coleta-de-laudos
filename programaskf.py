#coding: UTF-8

from platform import python_branch
from tkinter import Button
from selenium import webdriver                              #Importa a biblioteca Selenium
from webdriver_manager.chrome import ChromeDriverManager    #Importa a biblioteca que gerencia automaticamente a versão do chrome e selenium sem precisar instalar
from selenium.webdriver.chrome.service import Service       #Importa Service 
from selenium.webdriver import Chrome                       #Importa o Chrome
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl as xl    
import pandas as pd
import pyautogui                                            #Importa a biblioteca de automação
import datetime                                             #Importa a biblioteca data atual
import time                                                 #Importa a biblioteca time
import os                                                   
import glob
import os.path

def OpenWebSite():
    global navegador 
    #ERRO, NÃO ESTÁ TROCANDO O DIRETÓRIO DE DOWNLOAD       
    navegador = webdriver.Chrome("C:/Users/PC/Desktop/Teste automatização/SKF/driver/chromedriver.exe")                               #Abre o navegador
    navegador.get("https://repcenter.skf.com/machineviewer/logon.aspx")         #Entra no site SKF
    
    

        
    '''


    #DESATIVADO
    
    #ABRIR O SITE COM PYAUTOGUI
    
    
    
    pyautogui.moveTo(35,5)                                  #Move para a barra de tarefas
    pyautogui.click()                                       #Clica
    pyautogui.write("goog", interval=0.25)                  #Digita na barra de tarefas goog
    time.sleep(1)
    pyautogui.press("enter")                                #Pressiona ENTER
    time.sleep(10)
    pyautogui.moveTo(400,50)            
    pyautogui.click()                                       #Clica
    time.sleep(1)
    pyautogui.write("https://repcenter.skf.com/machineviewer/logon.aspx")     #Digita o site
    pyautogui.press("enter")            #Pressiona ENTER
    time.sleep(10)
    '''


def Login():
    
    login_path = '#identif_UserName'                                                    #VARIÁVEL PARA GUARDAR A CÉLULA DO LOGIN
    password_path = '#identif_Password'                                                 #VARIÁVEL PARA GUARDAR A CÉLULA DA SENHA

    login_element = navegador.find_element_by_css_selector(login_path)                  #ACESSA A CÉLULA DO LOGIN
    password_element = navegador.find_element_by_css_selector(password_path)            #ACESSA A CÉLULA DA SENHA

    login_element.send_keys('thomaz.silva')                                             #DIGITA O LOGIN
    password_element.send_keys('CXvm.r/$A[$Q>65E')                                      #DIGITA A SENHA
    
    navegador.find_element_by_xpath('//*[@id="identif_LoginButton"]').click()           #CLICA PARA LOGAR
    
    navegador.implicitly_wait(25)                                                                     #ESPERA 25 SEGUNDOS

    
 

    
    
    
    '''
    #FAZENDO LOGIN COM PYAUTOGUI (ANTIGO)
    pyautogui.moveTo(1000,500)                              #Move para o local do login
    pyautogui.click()                                       #Clica
    pyautogui.write("thomaz.silva",interval=0.10)           #Digita o login com intervalo
    time.sleep(2)
    pyautogui.press("tab")                                  #Muda para digitar a senha
    time.sleep(2)
    pyautogui.write("CXvm.r/$A[$Q>65E",interval=0.10)       #Escreve a senha
    pyautogui.press("enter")                                #Pressiona ENTER  
    time.sleep(20)
    '''

def GetDate():
    global hoje                                                                                         #TORNA HOJE UMA VARIÁVEL GLOBAL
    hoje = datetime.date.today()                                                                        #VARIÁVEL QUE HOJE (CONTÉM A DATA ATUAL)
    print(hoje)

def TreatDate():
    global treated_date
    
    treated_date = hoje.strftime('%d/%m/%Y')                                              #TRATA A DATA COMO DEVE SER ESCRITA 
    return treated_date


def Get_Worksheet():                                                                                 

    navegador.find_element_by_xpath('//*[@id="Reports"]/span[1]').click()    #CLICA NA CÉLULA RELATÓRIOS
    navegador.implicitly_wait(30) 
    
     
    navegador.find_element_by_css_selector("button[class ='MuiButtonBase-root MuiButton-root MuiButton-outlined MuiButton-outlinedPrimary']").click()               #FECHA A CAIXA DE SPAM 
    
    
    
    navegador.implicitly_wait(30)                                               # É UM TIME.SLEEP QUE SE ENCONTRAR O ELEMENTO EXECUTA ANTES
    navegador.find_element_by_xpath('//*[@id="detailedAssetHealth"]').click()                           #CLICA NA CÉLULA SAÚDE DETALHADA DO ATIVO
    navegador.implicitly_wait(30)  
    navegador.find_element_by_xpath('//*[@id="panel1a-content"]/div[1]/form/div[2]/div[1]/div/select').click()  #SELECIONA A HIERARQUIA
    navegador.implicitly_wait(30)  
    navegador.find_element_by_xpath('//*[@id="panel1a-content"]/div[1]/form/div[2]/div[1]/div/select/option[2]').click()  #SELECIONA BUNGE RIO GRANDE
    navegador.implicitly_wait(30)  
    

    initial_date_path = '#startDate'                                                      #VARIÁVEL PARA GUARDAR A CÉLULA DE INSERIR DATA INICIAL 
    #final_date_path = '#endDate'                                                         #VARIÁVEL PARA GUARDAR A CÉLULA DE INSERIR DAT'A FINAL
    
    initial_date_element = navegador.find_element_by_css_selector(initial_date_path)
    #final_date_element = navegador.find_element_by_css_selector(final_date_path)
                                                         
    #Não usado 
    # final_date_element = navegador.find_elements_by_css_selector(final_date_path)           #ACESSA A CÉLULA DATA FINAL


    navegador.implicitly_wait(30)
    initial_date_element.send_keys(Keys.CONTROL + "a")                                        #DA UM CONTROL A PARA SELECIONAR TODA A LINHA             
    initial_date_element.send_keys(Keys.DELETE)                                               #DA UM DELETE E LIMPA A LINHA
    initial_date_element.send_keys("29/01/2020")                                              #ESCREVER A DATA INICIAL  

    
    navegador.find_element_by_xpath('//*[@id="panel1a-content"]/div[1]/form/div[4]/div[3]/div/div[2]/div/label/span[2]').click()   #DESMARCA A CAIXA CONCLUÍDOS
    navegador.find_element_by_xpath('//*[@id="panel1a-content"]/div[2]/div/button/span[1]').click()         #CLICA NO LISTA DETALHADA DE ATIVOS

    

    #final_date_element.clear()
    #final_date_element.send_keys(hoje)                                                        #ESCREVER A DATA FINAL
    
    


    #pyautogui.moveTo(1120,350)
    #pyautogui.click()
    #pyautogui.scroll(+10000)    

    
    navegador.implicitly_wait(100)
    navegador.maximize_window()
    navegador.implicitly_wait(100)
    navegador.execute_script("window.scrollTo(0, -250)")                                            #Scrolla a página para funcionar o click no botão
    time.sleep(15)
    navegador.find_element_by_xpath('//*[@id="panel1a-content"]/div/div/div[1]/div[1]/button/span[1]').click()          #Clica nobotão para efetuar o download
    
    
    
    
    time.sleep(10)
     
    
     
     
     
                                                                  #SCROLLA A PÁGINA PRA CIMA
    #pyautogui.moveTo(1120,430)
    #pyautogui.click()
    
    
    
   
 



    '''
    #EXTRAIR COM PYAUTOGUI(ANTIGO)
    pyautogui.moveTo(100,350)                               #Move para a aba relatórios
    pyautogui.click()                                       #Clica
    time.sleep(1)
    pyautogui.moveTo(480,180)                               #Seleciona a aba Saúde detalhada do ativo
    pyautogui.click()                                       #Clica
    time.sleep(1)
    
    pyautogui.moveTo(500,450)                               #Seleciona a caixa
    pyautogui.click()                                       #Clica
    time.sleep(1)
    pyautogui.moveTo(500,500)                               #Seleciona a Bunge
    pyautogui.click()                                       #Clica
    
    pyautogui.moveTo(1340,680)                              #Seleciona a última caixa
    pyautogui.click()                                       #Clica
    '''
    

def get_download_path():
    """Returns the default downloads path for linux or windows"""
    if os.name == 'nt':
        import winreg
        sub_key = r'SOFTWARE\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
        downloads_guid = '{374DE290-123F-4565-9164-39C4925E467B}'
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, sub_key) as key:
            location = winreg.QueryValueEx(key, downloads_guid)[0]
        return location
    return os.path.join(os.path.expanduser('~'), 'downloads')

def get_latest_file_path(path):
    """return the more recent file in a path"""

    file_paths = glob.glob(f'{path}/*') # obtem o path de cada arquivo na pasta
    all_files_modification_time = [ os.path.getmtime(path) for path in file_paths ] # Obtem o modification time de cada arquivo
    latest_file_index = all_files_modification_time.index(max(all_files_modification_time)) # obtem o index do maior deles, o arquivo mais recente
    return file_paths[latest_file_index]


def Spreadsheet_Exchange():
    
    caminho = "C://Users/PC/Downloads"
    lista_arquivos = os.listdir(caminho)
    lista_datas = []

    for arquivo in lista_arquivos:
        # descobrir a data desse arquivo
        if ".xlsx" in arquivo:
            data = os.path.getmtime(f"{caminho}/{arquivo}")
            lista_datas.append((data, arquivo))
        
        # data inicial = 01/01/2021
        # data1 = 02/01/2021 -> 10.000
        # data2 = 15/02/2021 -> 150.000
        
    lista_datas.sort(reverse=True)
    ultimo_arquivo = lista_datas[0]
    print(ultimo_arquivo[1])
    '''
    #Filtrando o último arquivo excel adicionado na pasta downloads 
    pasta = "C:/Users/PC/Downloads"

    file_type = '\*xlsx'# se nao quiser filtrar por extenção deixe apenas *
    files = glob.glob(pasta + file_type)
    max_file = max(files, key=os.path.getctime) 
    for element in range(0, len(max_file)):
        print(string_name[element])
        
    '''
        
    #DESCOBRIR A DATA DO ARQUIVO

    ########################################################################################################################


    #Pegando o último arquivo instalado na pasta

    #data = os.path.getmtime(f"{pasta}/{arquivo}")
        

    '''
    lista_datas.sort(reverse=True)
    ultimo_arquivo = lista_datas[0]
    NomeDoArquivo = ultimo_arquivo[1]
    print(ultimo_arquivo[1])
    '''
    ########################################################################################################################





    #dados_planilha1 = pd.read_excel("C://Users/PC/Desktop/SKF/Downloads/"+(NomeDoArquivo))       # PEGA OS DADOS DA PLANILHA PARA TRATAMENTO DE DADOS


    #######################################################################################################################

    
    filename = ("C:\\Users\\PC\\Downloads\\"+(ultimo_arquivo[1]))
    wb1 = xl.load_workbook(filename) 
    ws1 = wb1.worksheets[0] 
    filename1 ="C:\\Users\\PC\\OneDrive - BUNGE\\Planejamento e Controle de Manutenção\\00. Confiabilidade\\05. Manutenção Preditiva - Vibração\\Laudos MHV.xlsx"
    wb2 = xl.load_workbook(filename1) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column 
    for i in range (1, mr + 300): 
        for j in range (1, mc + 1): 
            c = ws1.cell(row = i, column = j) 
            ws2.cell(row = i, column = j).value = c.value 
    wb2.save(str(filename1)) 
    





    #pyautogui.click(10,10)
    #time.sleep(2)
    #pyautogui.write('detailed')
    #time.sleep(2)
    #pyautogui.press('enter')
    #time.sleep(2)
    #pyautogui.press('left')
    #time.sleep(2)
    #pyautogui.press('enter')
    #time.sleep(2)
    #pyautogui.press('enter')
    #time.sleep(2)


    #--------------------------FAZ A CÓPIA DOS DADOS PARA OUTRA PLANILHA ------------------------------
    '''
    filename = (arquivo)
    wb1 = xl.load_workbook(filename) 
    ws1 = wb1.worksheets[0] 
    filename1 = "C:\\Users\\PC\\Desktop\\planilhateste.xlsx"
    wb2 = xl.load_workbook(filename1) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column 
    for i in range (1, mr + 1): 
        for j in range (1, mc + 1): 
            c = ws1.cell(row = i, column = j) 
            ws2.cell(row = i, column = j).value = c.value 
            wb2.save(str(filename1)) 

    '''


def ExtractInformation():
    GetDate()
    TreatDate()
    OpenWebSite()
    Login()
    Get_Worksheet()
    Spreadsheet_Exchange()
    print('Concluído')
    
    # Nome dos arquivos instalados começam com detailedAssetHealth

ExtractInformation()